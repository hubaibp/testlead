import time, re, csv, json, os, glob, threading, io
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from flask_socketio import SocketIO, emit
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch

app = Flask(__name__)
app.config['SECRET_KEY'] = 'leadgen_secret_2024'
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')








import os

# Render-specific settings
PORT = int(os.environ.get('PORT', 5000))
HOST = '0.0.0.0'

# In your run_scraper() function, update Chrome options:
opts = Options()
opts.add_argument("--headless=new")  # CRITICAL for Render!
opts.add_argument("--no-sandbox")
opts.add_argument("--disable-dev-shm-usage")
opts.add_argument("--disable-gpu")
opts.add_argument("--remote-debugging-port=9222")
opts.add_argument("--window-size=1920,1080")

# For WhatsApp Web (in init_whatsapp):
opts.add_argument("--headless=new")
opts.add_argument("--no-sandbox")
opts.add_argument("--disable-dev-shm-usage")























# Global state
scraper_instance = None
scraper_thread = None
all_businesses = []
no_website_leads = []
whatsapp_ready = False
scraping_active = False
whatsapp_thread = None

# ── Persistent Chrome profile path for WhatsApp (so QR is saved) ──
WA_PROFILE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "wa_chrome_profile")

def emit_log(msg, level="info"):
    socketio.emit('log', {'msg': msg, 'level': level})

def emit_stats():
    socketio.emit('stats', {
        'total': len(all_businesses),
        'leads': len(no_website_leads),
        'with_website': len(all_businesses) - len(no_website_leads)
    })

def emit_business(biz, category):
    socketio.emit('new_business', {'business': biz, 'category': category})

# ─── Scraper Logic ───────────────────────────────────────────────────────────

def is_real_website(url):
    if not url: return False
    fake = ['google.com','google.co.in','maps.google.com','facebook.com','fb.com',
            'instagram.com','twitter.com','x.com','linkedin.com','youtube.com',
            'justdial.com','sulekha.com','indiamart.com','wordpress.com','blogger.com',
            'wix.com','weebly.com']
    url_l = url.lower()
    for f in fake:
        if f in url_l: return False
    suspicious = ['/intl/en/about/','tab=lh','products?tab','maps/place/',
                  '/pages/','/profile','/feed','/reel','/post']
    for p in suspicious:
        if p in url_l: return False
    if url_l.startswith('http'):
        domain = url_l.split('//')[-1].split('/')[0]
        if '.' in domain: return True
    return False

def run_scraper(business_type, location, max_businesses):
    global scraper_instance, all_businesses, no_website_leads, scraping_active
    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC

        all_businesses = []
        no_website_leads = []
        scraping_active = True
        socketio.emit('scraping_started')

        opts = Options()
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option('useAutomationExtension', False)
        opts.add_argument("--start-maximized")
        opts.add_argument("--lang=en")

        driver = webdriver.Chrome(options=opts)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        wait = WebDriverWait(driver, 10)
        scraper_instance = driver

        emit_log(f"🔍 Searching: {business_type} in {location}", "info")

        query = f"{business_type} {location}"
        url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"
        driver.get(url)
        time.sleep(5)

        scrollable_div = None
        for selector in ["div[role='feed']", "div[style*='overflow-y']"]:
            try:
                scrollable_div = driver.find_element(By.CSS_SELECTOR, selector)
                if scrollable_div: break
            except: continue

        if not scrollable_div:
            emit_log("❌ Could not find results panel", "error")
            driver.quit()
            scraping_active = False
            socketio.emit('scraping_done')
            return

        seen_hrefs = set()
        businesses_raw = []
        last_count = 0
        no_new = 0

        emit_log("📜 Scrolling to collect all businesses first...", "info")
        while len(businesses_raw) < max_businesses and no_new < 12:
            driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", scrollable_div)
            time.sleep(3)
            elements = driver.find_elements(By.CSS_SELECTOR, "a[href*='https://www.google.com/maps/place']")
            for elem in elements:
                try:
                    name = elem.get_attribute("aria-label")
                    href = elem.get_attribute("href")
                    if name and href and href not in seen_hrefs:
                        seen_hrefs.add(href)
                        businesses_raw.append({'name': name, 'href': href})
                except: pass
            if len(businesses_raw) > last_count:
                emit_log(f"📍 Found {len(businesses_raw)} businesses so far...", "info")
                last_count = len(businesses_raw)
                no_new = 0
            else:
                no_new += 1
                emit_log(f"⏳ No new ({no_new}/12), still scrolling...", "info")
            try:
                for btn in driver.find_elements(By.XPATH, "//button[contains(.,'More results')]"):
                    if btn.is_displayed():
                        btn.click()
                        time.sleep(2)
                        no_new = 0
                        break
            except: pass

        businesses_raw = businesses_raw[:max_businesses]
        emit_log(f"✅ Collected {len(businesses_raw)} businesses. Extracting details...", "success")

        for i, biz in enumerate(businesses_raw):
            if not scraping_active: break
            emit_log(f"[{i+1}/{len(businesses_raw)}] Processing: {biz['name'][:40]}", "info")
            details = extract_details_by_href(driver, biz)
            if details:
                all_businesses.append(details)
                if not details['has_website'] and details['phone'] != 'Not available' and len(details['phone']) >= 10:
                    no_website_leads.append(details)
                    emit_business(details, 'lead')
                    emit_log(f"🎯 LEAD: {details['name']} | {details['phone']}", "success")
                else:
                    emit_business(details, 'all')
                emit_stats()
            time.sleep(2.5)

        driver.quit()
        scraping_active = False
        emit_log(f"🏁 Done! Total: {len(all_businesses)} | Leads: {len(no_website_leads)}", "success")
        socketio.emit('scraping_done', {'total': len(all_businesses), 'leads': len(no_website_leads)})

    except Exception as e:
        emit_log(f"❌ Error: {str(e)}", "error")
        scraping_active = False
        socketio.emit('scraping_done')

def extract_details_by_href(driver, biz):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    try:
        driver.get(biz['href'])
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "h1"))
            )
        except: pass
        time.sleep(2)

        phone = 'Not available'
        try:
            pe = WebDriverWait(driver, 6).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "button[data-item-id*='phone']"))
            )
            lbl = pe.get_attribute('aria-label') or ''
            clean = re.sub(r'[^0-9+]', '', lbl)
            phone = clean if clean and len(clean) >= 10 else 'Not available'
        except: pass

        address = 'Not available'
        pincode = 'Not available'
        try:
            ae = driver.find_element(By.CSS_SELECTOR, "button[data-item-id='address']")
            address = ae.get_attribute('aria-label') or 'Not available'
            m = re.search(r'\b\d{6}\b', address)
            pincode = m.group() if m else 'Not found'
        except: pass

        rating = 'No rating'
        try:
            re_ = driver.find_element(By.CSS_SELECTOR, "div[aria-label*='stars']")
            rt = re_.get_attribute('aria-label') or ''
            rm = re.search(r'([\d.]+)', rt)
            rating = rm.group(1) if rm else 'No rating'
        except: pass

        category = 'Not specified'
        try:
            ce = driver.find_element(By.CSS_SELECTOR, "button[data-item-id='category']")
            category = ce.text or 'Not specified'
        except:
            try:
                spans = driver.find_elements(By.CSS_SELECTOR, "span.DkEaL")
                if spans: category = spans[0].text or 'Not specified'
            except: pass

        has_website, website_url = False, 'NO WEBSITE'
        for sel in ["a[data-item-id='authority']", "a[data-tooltip='Open website']", "a[aria-label*='website' i]"]:
            try:
                for el in driver.find_elements(By.CSS_SELECTOR, sel):
                    href = el.get_attribute('href')
                    if href and is_real_website(href):
                        has_website = True
                        website_url = href
                        break
                if has_website: break
            except: pass

        return {
            'name': biz['name'],
            'address': address,
            'phone': phone,
            'website': website_url if has_website else 'NO WEBSITE',
            'has_website': has_website,
            'rating': rating,
            'category': category,
            'pincode': pincode,
            'scraped_at': datetime.now().strftime("%Y-%m-%d %H:%M")
        }
    except Exception as e:
        emit_log(f"⚠ Extract error: {str(e)[:60]}", "warning")
        return None

# ─── WhatsApp Logic ──────────────────────────────────────────────────────────

wa_driver = None
wa_ready = False

def init_whatsapp():
    """
    Opens WhatsApp Web using a PERSISTENT Chrome profile so the user only
    needs to scan QR once.  Profile is stored in ./wa_chrome_profile/
    """
    global wa_driver, wa_ready
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.options import Options

    opts = Options()
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("detach", True)   # keep browser open even if Python crashes
    opts.add_argument("--start-maximized")
    opts.add_argument("--lang=en-US")

    # ── KEY FIX: persistent profile so login is remembered ──
    os.makedirs(WA_PROFILE_DIR, exist_ok=True)
    opts.add_argument(f"--user-data-dir={WA_PROFILE_DIR}")
    opts.add_argument("--profile-directory=Default")

    wa_driver = webdriver.Chrome(options=opts)
    wa_driver.get("https://web.whatsapp.com/")

    emit_log("📱 WhatsApp Web opening... waiting for login.", "info")
    socketio.emit('whatsapp_qr')

    # Wait up to 120 seconds for chat list (will skip QR if already logged in)
    for i in range(120):
        try:
            wa_driver.find_element(By.CSS_SELECTOR, "[data-testid='chat-list'], #side, ._aigs")
            wa_ready = True
            emit_log("✅ WhatsApp connected!", "success")
            socketio.emit('whatsapp_ready')
            return True
        except:
            time.sleep(1)
            if i % 20 == 0 and i > 0:
                emit_log(f"⏳ Waiting for WhatsApp login... {120-i}s left", "info")

    emit_log("❌ WhatsApp login timed out", "error")
    socketio.emit('whatsapp_failed')
    return False



def send_wa_messages(leads, message_template, max_msgs, delay_sec):
    """
    Sends messages using clipboard paste — most reliable method for
    WhatsApp Web contenteditable boxes.
    """
    global wa_driver, wa_ready
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.action_chains import ActionChains  # FIXED IMPORT
    import pyperclip   # pip install pyperclip

    if not wa_ready or not wa_driver:
        emit_log("❌ WhatsApp not ready", "error")
        return

    sent = 0
    failed = 0

    # Selectors for the message input box (ordered by reliability)
    BOX_SELECTORS = [
        (By.CSS_SELECTOR, "div[data-testid='conversation-compose-box-input']"),
        (By.CSS_SELECTOR, "div[contenteditable='true'][data-tab='10']"),
        (By.CSS_SELECTOR, "div[contenteditable='true'][data-tab='6']"),
        (By.CSS_SELECTOR, "footer div[contenteditable='true']"),
        (By.XPATH,        "//footer//div[@contenteditable='true']"),
        (By.XPATH,        "//div[@contenteditable='true' and @role='textbox']"),
    ]

    for i, lead in enumerate(leads[:max_msgs]):
        phone = str(lead.get('phone', '')).strip()
        phone = ''.join(filter(str.isdigit, phone))

        if len(phone) < 10:
            emit_log(f"⚠ Invalid phone: {lead.get('name')}", "warning")
            failed += 1
            continue

        # Normalise to international format (India default)
        if len(phone) == 10:
            phone = f"91{phone}"
        elif not phone.startswith('91') and len(phone) < 12:
            phone = f"91{phone[-10:]}"

        name = lead.get('name', 'Business')
        msg  = message_template.replace('{name}', name).replace('{business_name}', name)

        try:
            # Navigate directly via wa.me deep-link — most reliable opener
            deep_link = f"https://web.whatsapp.com/send?phone={phone}&text="
            wa_driver.get(deep_link)
            emit_log(f"📲 Opening chat for {name} ({phone})...", "info")
            time.sleep(10)   # give WA Web time to load the chat

            # ── Check for "phone number not on WhatsApp" popup ──
            try:
                popup = wa_driver.find_elements(
                    By.XPATH,
                    "//div[@data-testid='popup-contents'] | //div[contains(@class,'_aiku')]"
                )
                if popup:
                    # click OK / close button
                    for btn_xpath in [
                        "//div[@data-testid='popup-contents']//div[@role='button']",
                        "//button[.//div[text()='OK']]",
                        "//div[@role='button'][contains(.,'OK')]",
                    ]:
                        try:
                            wa_driver.find_element(By.XPATH, btn_xpath).click()
                            break
                        except: pass
                    emit_log(f"⚠ {name}: number not on WhatsApp or invalid", "warning")
                    failed += 1
                    continue
            except: pass

            # ── Find the message input box ──
            box = None
            for by, sel in BOX_SELECTORS:
                try:
                    el = WebDriverWait(wa_driver, 6).until(
                        EC.element_to_be_clickable((by, sel))
                    )
                    if el and el.is_displayed():
                        box = el
                        break
                except: pass

            if not box:
                emit_log(f"❌ Message box not found for {name}", "warning")
                failed += 1
                continue

            # ── PASTE METHOD: most reliable on modern WhatsApp Web ──
            try:
                pyperclip.copy(msg)
                box.click()
                time.sleep(0.4)
                # Ctrl+V paste
                ActionChains(wa_driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                time.sleep(0.8)
            except Exception as paste_err:
                # Fallback: JS clipboard API
                try:
                    wa_driver.execute_script("""
                        const dt = new DataTransfer();
                        dt.setData('text/plain', arguments[0]);
                        arguments[1].focus();
                        arguments[1].dispatchEvent(new ClipboardEvent('paste', {clipboardData: dt, bubbles: true}));
                    """, msg, box)
                    time.sleep(0.8)
                except:
                    emit_log(f"❌ Could not type message for {name}: {str(paste_err)[:60]}", "error")
                    failed += 1
                    continue

            # ── Send ──
            ActionChains(wa_driver).send_keys(Keys.ENTER).perform()
            time.sleep(1.5)

            sent += 1
            emit_log(f"✅ [{i+1}/{min(len(leads), max_msgs)}] Sent → {name} ({phone})", "success")
            socketio.emit('wa_progress', {
                'sent': sent,
                'failed': failed,
                'total': min(len(leads), max_msgs)
            })

            # ── Delay between messages (anti-ban) ──
            if i < min(len(leads), max_msgs) - 1:
                remaining = delay_sec
                while remaining > 0:
                    time.sleep(min(5, remaining))
                    remaining -= 5
                    if remaining > 0:
                        emit_log(f"⏱ Next message in {remaining}s...", "info")

        except Exception as e:
            emit_log(f"❌ Failed: {name} — {str(e)[:80]}", "error")
            failed += 1

    emit_log(f"🏁 WhatsApp done! Sent: {sent} | Failed: {failed}", "success")
    socketio.emit('wa_done', {'sent': sent, 'failed': failed})


# ─── Routes ──────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/scrape', methods=['POST'])
def start_scrape():
    global scraper_thread, scraping_active
    if scraping_active:
        return jsonify({'error': 'Scraping already in progress'}), 400
    data = request.json
    btype    = data.get('business_type', 'businesses')
    location = data.get('location', '')
    max_b    = min(int(data.get('max_businesses', 50)), 200)
    scraper_thread = threading.Thread(
        target=run_scraper, args=(btype, location, max_b), daemon=True
    )
    scraper_thread.start()
    return jsonify({'status': 'started'})

@app.route('/api/stop_scrape', methods=['POST'])
def stop_scrape():
    global scraping_active
    scraping_active = False
    emit_log("⛔ Scraping stopped by user", "warning")
    return jsonify({'status': 'stopped'})

@app.route('/api/clear_data', methods=['POST'])
def clear_data():
    """Clear all scraped data from memory."""
    global all_businesses, no_website_leads
    all_businesses   = []
    no_website_leads = []
    socketio.emit('data_cleared')
    return jsonify({'status': 'cleared'})

@app.route('/api/businesses')
def get_businesses():
    btype = request.args.get('type', 'all')
    if btype == 'leads':
        return jsonify(no_website_leads)
    return jsonify(all_businesses)

@app.route('/api/whatsapp/init', methods=['POST'])
def start_whatsapp():
    t = threading.Thread(target=init_whatsapp, daemon=True)
    t.start()
    return jsonify({'status': 'initializing'})

@app.route('/api/whatsapp/send', methods=['POST'])
def send_whatsapp():
    data        = request.json
    use_leads   = data.get('use_leads', True)
    source      = no_website_leads if use_leads else all_businesses
    msg_template = data.get('message', 'Hi {name}, I can help build your website!')
    max_msgs    = min(int(data.get('max_messages', 20)), 50)
    delay       = max(int(data.get('delay', 25)), 20)
    t = threading.Thread(
        target=send_wa_messages,
        args=(source, msg_template, max_msgs, delay),
        daemon=True
    )
    t.start()
    return jsonify({'status': 'sending', 'total': min(len(source), max_msgs)})

@app.route('/api/export/excel')
def export_excel():
    dtype = request.args.get('type', 'leads')
    data  = no_website_leads if dtype == 'leads' else all_businesses
    if not data:
        return jsonify({'error': 'No data'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads" if dtype == 'leads' else "All Businesses"

    headers = ['Name', 'Phone', 'Address', 'Rating', 'Category', 'Pincode', 'Website', 'Scraped At']
    keys    = ['name', 'phone', 'address', 'rating', 'category', 'pincode', 'website', 'scraped_at']

    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="00ff88", bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    alt_fill = PatternFill(start_color="f0f4ff", end_color="f0f4ff", fill_type="solid")
    for row, biz in enumerate(data, 2):
        for col, key in enumerate(keys, 1):
            cell = ws.cell(row=row, column=col, value=biz.get(key, ''))
            if row % 2 == 0:
                cell.fill = alt_fill
            ws.column_dimensions[cell.column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{'leads' if dtype=='leads' else 'all_businesses'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname
    )

@app.route('/api/export/pdf')
def export_pdf():
    dtype = request.args.get('type', 'leads')
    data  = no_website_leads if dtype == 'leads' else all_businesses
    if not data:
        return jsonify({'error': 'No data'}), 400

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elems  = []

    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=16,
                                 textColor=colors.HexColor('#1a1a2e'), spaceAfter=12)
    elems.append(Paragraph(
        f"{'Leads (No Website)' if dtype=='leads' else 'All Businesses'} — {datetime.now().strftime('%Y-%m-%d')}",
        title_style
    ))
    elems.append(Paragraph(f"Total: {len(data)} records", styles['Normal']))
    elems.append(Spacer(1, 12))

    headers    = ['Name', 'Phone', 'Rating', 'Category', 'Pincode']
    keys       = ['name', 'phone', 'rating', 'category', 'pincode']
    table_data = [headers]
    for biz in data:
        table_data.append([str(biz.get(k, ''))[:30] for k in keys])

    t = Table(table_data, colWidths=[2.2*inch, 1.4*inch, 0.8*inch, 1.4*inch, 0.8*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1,  0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR',     (0, 0), (-1,  0), colors.HexColor('#00ff88')),
        ('FONTNAME',      (0, 0), (-1,  0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1,  0), 10),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4ff')]),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('GRID',          (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING',       (0, 0), (-1, -1), 4),
    ]))
    elems.append(t)
    doc.build(elems)
    buf.seek(0)
    fname = f"{'leads' if dtype=='leads' else 'all_businesses'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=fname)

@app.route('/api/export/csv')
def export_csv():
    dtype = request.args.get('type', 'leads')
    data  = no_website_leads if dtype == 'leads' else all_businesses
    if not data:
        return jsonify({'error': 'No data'}), 400
    buf  = io.StringIO()
    keys = ['name', 'phone', 'address', 'rating', 'category', 'pincode', 'website', 'scraped_at']
    w    = csv.DictWriter(buf, fieldnames=keys, extrasaction='ignore')
    w.writeheader()
    w.writerows(data)
    buf.seek(0)
    fname = f"{'leads' if dtype=='leads' else 'businesses'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(
        io.BytesIO(buf.getvalue().encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name=fname
    )

@app.route('/api/status')
def get_status():
    return jsonify({
        'scraping': scraping_active,
        'wa_ready': wa_ready,
        'total':    len(all_businesses),
        'leads':    len(no_website_leads)
    })

if __name__ == '__main__':
    socketio.run(app, debug=False, host='0.0.0.0', port=5000, allow_unsafe_werkzeug=True)
    
if __name__ == '__main__':
    socketio.run(app, debug=False, host=HOST, port=PORT, allow_unsafe_werkzeug=True)    